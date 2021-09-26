import openpyxl as xl
from root import main_fun
def origin():
    try:
      directory = input("ENTER THE PATH OF FILE IN PROPER FORMAT\n")
      wb = xl.load_workbook(f"{directory}")
      sheets=wb.sheetnames
    except FileNotFoundError:
        print("\nFILE NOT FOUND PLEASE TYPE NAME CORRECTLY\n")
        origin()
    try:
       print(f"SHEETS AVAILABLE {sheets} FOR U")
       print("\n")
       sheet = input("ENTER THE NAME OF SHEET ON WHICH YOU WANT TO PERFORM OPERATION\n")
       s = wb[f'{sheet}']
    except KeyError:
        print("\nSHEET NOT FOUND PLEASE TYPE NAME OF FILE AND SHEET NAME CORRECTLY\n")
        origin()
    row = s.max_row
    column = s.max_column
    main_fun(s, row, column,wb)
origin()

